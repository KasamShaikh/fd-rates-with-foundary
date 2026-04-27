"""
Local development server — Runs the same API logic as Azure Functions
using Flask. No Azure Functions Core Tools required.

Usage: python dev_server.py
"""

import json
import logging
import os
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, request, jsonify
from flask_cors import CORS
from azure.storage.blob import BlobServiceClient, ContentSettings
from azure.identity import DefaultAzureCredential

# Load .env from project root (one level up from backend/)
_env_path = Path(__file__).resolve().parent.parent / ".env"
if _env_path.exists():
    load_dotenv(_env_path)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Allowed CORS origins. Defaults to the local dev server. In cloud, set
# ALLOWED_ORIGINS to a comma-separated list e.g.
#   ALLOWED_ORIGINS=https://my-swa.azurestaticapps.net
_default_origins = "http://localhost:3000"
_origins_csv = os.environ.get("ALLOWED_ORIGINS", _default_origins)
_origins = [o.strip() for o in _origins_csv.split(",") if o.strip()]
CORS(app, origins=_origins)

URLS_FILE = os.path.join(os.path.dirname(__file__), "urls.json")
BLOB_CONTAINER = os.environ.get("BLOB_CONTAINER_NAME", "fd-rates")

# Local file-based storage for dev. Disable in cloud by setting
# LOCAL_RESULTS_ENABLED=false so the container only writes to Blob Storage.
LOCAL_RESULTS_ENABLED = (
    os.environ.get("LOCAL_RESULTS_ENABLED", "true").lower() != "false"
)
LOCAL_RESULTS_DIR = os.path.join(os.path.dirname(__file__), "_local_results")
if LOCAL_RESULTS_ENABLED:
    os.makedirs(LOCAL_RESULTS_DIR, exist_ok=True)


def _get_blob_service_client():
    """Return a BlobServiceClient using DefaultAzureCredential (no connection string needed)."""
    account_name = os.environ.get("STORAGE_ACCOUNT_NAME", "")
    if not account_name:
        return None
    account_url = f"https://{account_name}.blob.core.windows.net"
    return BlobServiceClient(
        account_url=account_url, credential=DefaultAzureCredential()
    )


def _upload_to_blob(
    blob_name: str, data: bytes, content_type: str = "application/json"
) -> bool:
    """Upload bytes to Azure Blob Storage. Returns True on success, False on failure."""
    try:
        client = _get_blob_service_client()
        if client is None:
            logger.warning("STORAGE_ACCOUNT_NAME not set — skipping blob upload")
            return False
        container_client = client.get_container_client(BLOB_CONTAINER)
        container_client.upload_blob(
            name=blob_name,
            data=data,
            overwrite=True,
            content_settings=ContentSettings(content_type=content_type),
        )
        logger.info("Uploaded to blob: %s/%s", BLOB_CONTAINER, blob_name)
        return True
    except Exception as e:
        logger.warning("Blob upload failed for %s: %s", blob_name, e)
        return False


URLS_BLOB_NAME = "urls.json"


def _download_urls_from_blob():
    """Return parsed urls list from blob, or None if not found / unavailable."""
    try:
        client = _get_blob_service_client()
        if client is None:
            return None
        blob_client = client.get_blob_client(BLOB_CONTAINER, URLS_BLOB_NAME)
        if not blob_client.exists():
            return None
        data = blob_client.download_blob().readall()
        return json.loads(data.decode("utf-8"))
    except Exception as e:
        logger.warning("Failed to download urls.json from blob: %s", e)
        return None


def _upload_urls_to_blob(urls) -> bool:
    payload = json.dumps(urls, indent=2, ensure_ascii=False).encode("utf-8")
    return _upload_to_blob(URLS_BLOB_NAME, payload, content_type="application/json")


def _load_urls():
    """Load URL list. Blob is authoritative; falls back to bundled file on first run."""
    blob_urls = _download_urls_from_blob()
    if blob_urls is not None:
        return blob_urls
    # First-run seed: read the bundled urls.json and push it to blob so it
    # becomes the source of truth from this point on.
    if os.path.exists(URLS_FILE):
        with open(URLS_FILE, "r", encoding="utf-8") as f:
            seed = json.load(f)
        _upload_urls_to_blob(seed)
        return seed
    return []


def _save_urls(urls):
    # Authoritative copy goes to blob storage.
    _upload_urls_to_blob(urls)
    # Best-effort local mirror (useful for dev; ephemeral in cloud).
    try:
        with open(URLS_FILE, "w", encoding="utf-8") as f:
            json.dump(urls, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.warning("Failed to write local urls.json mirror: %s", e)


def _save_local_result(filename, data):
    """Persist a result file to the local cache. No-op when LOCAL_RESULTS_ENABLED=false (cloud)."""
    if not LOCAL_RESULTS_ENABLED:
        return
    path = os.path.join(LOCAL_RESULTS_DIR, filename)
    if isinstance(data, (dict, list)):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str)
    else:
        with open(path, "wb") as f:
            f.write(data)


def _load_local_result(filename):
    """Return a saved result. Reads local cache first; falls back to Blob Storage when disabled or missing."""
    if LOCAL_RESULTS_ENABLED:
        path = os.path.join(LOCAL_RESULTS_DIR, filename)
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    # Cloud / missing-local fallback: pull from blob
    try:
        client = _get_blob_service_client()
        if client is None:
            return None
        blob = client.get_container_client(BLOB_CONTAINER).get_blob_client(filename)
        if not blob.exists():
            return None
        return json.loads(blob.download_blob().readall().decode("utf-8"))
    except Exception as e:
        logger.info("Blob load fallback for %s failed: %s", filename, e)
        return None


# -------------------------------------------------------
# GET /api/urls
# -------------------------------------------------------
@app.route("/api/urls", methods=["GET"])
def list_urls():
    return jsonify(_load_urls())


# -------------------------------------------------------
# POST /api/urls
# -------------------------------------------------------
@app.route("/api/urls", methods=["POST"])
def add_url():
    body = request.get_json(silent=True) or {}
    url = body.get("url", "").strip()
    bank_name = body.get("bank_name", "").strip()

    if not url or not bank_name:
        return jsonify({"error": "Both 'url' and 'bank_name' are required"}), 400

    urls = _load_urls()
    new_entry = {
        "id": str(uuid.uuid4()),
        "url": url,
        "bank_name": bank_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    urls.append(new_entry)
    _save_urls(urls)
    return jsonify(new_entry), 201


# -------------------------------------------------------
# DELETE /api/urls/<id>
# -------------------------------------------------------
@app.route("/api/urls/<url_id>", methods=["DELETE"])
def delete_url(url_id):
    urls = _load_urls()
    original_len = len(urls)
    urls = [u for u in urls if u["id"] != url_id]
    if len(urls) == original_len:
        return jsonify({"error": "URL not found"}), 404
    _save_urls(urls)
    return jsonify({"message": "Deleted", "id": url_id})


# -------------------------------------------------------
# PUT /api/urls/<id>  — edit bank_name and/or url
# -------------------------------------------------------
@app.route("/api/urls/<url_id>", methods=["PUT"])
def update_url(url_id):
    body = request.get_json(silent=True) or {}
    new_url = (body.get("url") or "").strip()
    new_name = (body.get("bank_name") or "").strip()
    if not new_url and not new_name:
        return jsonify({"error": "Provide 'url' and/or 'bank_name'"}), 400
    urls = _load_urls()
    found = None
    for u in urls:
        if u["id"] == url_id:
            if new_url:
                u["url"] = new_url
            if new_name:
                u["bank_name"] = new_name
            u["updated_at"] = datetime.now(timezone.utc).isoformat()
            found = u
            break
    if found is None:
        return jsonify({"error": "URL not found"}), 404
    _save_urls(urls)
    return jsonify(found)


# -------------------------------------------------------
# POST /api/scrape  (fire-and-forget)
# -------------------------------------------------------
# Tracks the most recent background run so duplicate POSTs don't double-start.
_scrape_thread_lock = threading.Lock()
_scrape_thread: "threading.Thread | None" = None


def _run_scrape_job(urls, force_refresh: bool) -> None:
    """Background worker: runs the scrape, persists results, never raises."""
    from agent.fd_rate_agent import scrape_all_urls
    from agent import progress as _progress
    import time as _time

    if force_refresh:
        os.environ["FORCE_REFRESH"] = "true"

    _t0 = _time.monotonic()
    elapsed_seconds = 0.0
    scrape_output = None
    try:
        scrape_output = scrape_all_urls(urls)
    except Exception as e:
        logger.exception("Fetch failed")
        _progress.log(f"Fetch failed: {e}", level="error")
        return
    finally:
        elapsed_seconds = round(_time.monotonic() - _t0, 1)
        mins, secs = divmod(int(elapsed_seconds), 60)
        _progress.log(
            f"Total time taken: {mins}m {secs}s ({elapsed_seconds:.1f} seconds).",
            level="success",
        )
        if force_refresh:
            os.environ.pop("FORCE_REFRESH", None)

    try:
        results = (
            scrape_output.get("results", scrape_output)
            if isinstance(scrape_output, dict)
            else scrape_output
        )
        token_usage = (
            scrape_output.get("token_usage", {})
            if isinstance(scrape_output, dict)
            else {}
        )
        di_pages = (
            scrape_output.get("di_pages", 0) if isinstance(scrape_output, dict) else 0
        )
        unchanged_count = (
            scrape_output.get("unchanged_count", 0)
            if isinstance(scrape_output, dict)
            else 0
        )

        timestamp = datetime.now(timezone.utc)
        payload = {
            "scraped_at": timestamp.isoformat(),
            "bank_count": len(results) if results is not None else 0,
            "token_usage": token_usage,
            "di_pages": di_pages,
            "unchanged_count": unchanged_count,
            "elapsed_seconds": elapsed_seconds,
            "results": results or [],
        }

        ts_str = timestamp.strftime("%Y%m%d_%H%M%S")
        _save_local_result(f"fd_rates_{ts_str}.json", payload)
        _save_local_result("latest.json", payload)

        payload_bytes = json.dumps(payload, ensure_ascii=False, default=str).encode(
            "utf-8"
        )
        _upload_to_blob(f"fd_rates_{ts_str}.json", payload_bytes, "application/json")
        _upload_to_blob("latest.json", payload_bytes, "application/json")

        logger.info("Fetch complete.")
    except Exception:
        logger.exception("Failed to persist scrape results")
        _progress.log("Failed to persist results", level="error")
    finally:
        _progress.mark_done()


@app.route("/api/scrape", methods=["POST"])
def scrape_all():
    global _scrape_thread

    urls = _load_urls()
    if not urls:
        return jsonify({"error": "No URLs configured. Add URLs first."}), 400

    # Optional: filter by selected URL ids in request body. Default = scrape all.
    body = request.get_json(silent=True) or {}
    selected_ids = body.get("ids")
    if isinstance(selected_ids, list) and len(selected_ids) > 0:
        wanted = {str(i) for i in selected_ids}
        urls = [u for u in urls if str(u.get("id")) in wanted]
        if not urls:
            return jsonify(
                {"error": "None of the selected URL ids match configured URLs."}
            ), 400

    force_refresh = bool(body.get("force"))

    # Validate that Foundry is configured
    if not os.environ.get("PROJECT_ENDPOINT") or "<your-resource>" in os.environ.get(
        "PROJECT_ENDPOINT", ""
    ):
        return jsonify(
            {
                "error": "PROJECT_ENDPOINT not configured. Run setup.ps1 first or set it in .env file.",
                "hint": "Copy .env.template to .env and fill in your Azure AI Foundry project endpoint.",
            }
        ), 503

    from agent import progress as _progress

    with _scrape_thread_lock:
        if _scrape_thread is not None and _scrape_thread.is_alive():
            snap = _progress.snapshot(since=0)
            return jsonify(
                {
                    "started": False,
                    "already_running": True,
                    "run_id": snap.get("run_id"),
                    "bank_count": len(urls),
                }
            ), 409

        run_id = _progress.reset()
        logger.info("Starting fetch for %d URLs (run_id=%s)", len(urls), run_id)
        t = threading.Thread(
            target=_run_scrape_job,
            args=(urls, force_refresh),
            name=f"scrape-run-{run_id}",
            daemon=True,
        )
        _scrape_thread = t
        t.start()

    return jsonify(
        {
            "started": True,
            "run_id": run_id,
            "bank_count": len(urls),
        }
    ), 202


# -------------------------------------------------------
# GET /api/scrape/progress
# -------------------------------------------------------
@app.route("/api/scrape/progress", methods=["GET"])
def scrape_progress():
    """Return live progress events for the current/last scrape run.

    Query params:
      since (int, optional) — return events with index >= since (for incremental polling)
    """
    from agent import progress as _progress

    try:
        since = int(request.args.get("since", "0"))
    except ValueError:
        since = 0
    return jsonify(_progress.snapshot(since=since))


# -------------------------------------------------------
# POST /api/scrape/cancel — request graceful stop of in-flight run
# -------------------------------------------------------
@app.route("/api/scrape/cancel", methods=["POST"])
def scrape_cancel():
    """Flag the current scrape run for cancellation.

    Workers check this flag between banks (and between agent-poll ticks) and
    bail out early. The currently-running bank's agent run is asked to cancel
    too, so token spend stops as quickly as Foundry will allow.
    """
    from agent import progress as _progress

    requested = _progress.cancel()
    if requested:
        _progress.log(
            "⏹️ Stop requested — cancelling current bank and skipping the rest.",
            level="warn",
        )
    return jsonify({"cancelled": requested})


# -------------------------------------------------------
# GET /api/results/latest
# -------------------------------------------------------
@app.route("/api/results/latest", methods=["GET"])
def get_latest_results():
    data = _load_local_result("latest.json")
    if data is None:
        return jsonify({"error": "No results found. Run a scrape first."}), 404
    return jsonify(data)


# -------------------------------------------------------
# DELETE /api/results/latest — used by Reset Screen
# -------------------------------------------------------
@app.route("/api/results/latest", methods=["DELETE"])
def delete_latest_results():
    removed_local = False
    if LOCAL_RESULTS_ENABLED:
        path = os.path.join(LOCAL_RESULTS_DIR, "latest.json")
        if os.path.exists(path):
            try:
                os.remove(path)
                removed_local = True
            except OSError as e:
                logger.warning("Could not delete %s: %s", path, e)

    removed_blob = False
    try:
        client = _get_blob_service_client()
        if client is not None:
            container_client = client.get_container_client(BLOB_CONTAINER)
            container_client.delete_blob("latest.json")
            removed_blob = True
    except Exception as e:
        # Treat "not found" as success
        logger.info("Blob latest.json delete: %s", e)

    return jsonify(
        {
            "message": "Latest result cleared",
            "removed_local": removed_local,
            "removed_blob": removed_blob,
        }
    )


# -------------------------------------------------------
# POST /api/export-excel
# -------------------------------------------------------
@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = _load_local_result("latest.json")
    if data is None:
        return jsonify({"error": "No results found to export"}), 404

    results = data.get("results", [])
    if not results:
        return jsonify({"error": "No bank results to export"}), 400

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for bank_result in results:
        bank_name = bank_result.get("bank_name", "Unknown")
        # Excel sheet titles cannot contain: \ / ? * [ ] : and must be <= 31 chars
        safe_title = (
            "".join(c if c not in "\\/?*[]:" else "-" for c in bank_name)[:31].strip()
            or "Unknown"
        )
        # Ensure uniqueness (openpyxl raises on duplicate sheet names)
        base_title = safe_title
        suffix = 2
        while safe_title in wb.sheetnames:
            safe_title = f"{base_title[:28]}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(title=safe_title)

        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{bank_name} — Fixed Deposit Rates"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")

        eff_date = bank_result.get("effective_date", "N/A")
        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Effective Date: {eff_date}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        if "error" in bank_result:
            ws["A4"].value = f"Error: {bank_result.get('error')}"
            ws["A5"].value = f"Reason: {bank_result.get('reason', 'N/A')}"
            continue

        headers = [
            "Category",
            "Amount Slab",
            "Scheme",
            "Tenor",
            "Min Days",
            "Max Days",
            "Rate (%)",
            "Additional Info",
        ]
        col_widths = [20, 20, 18, 25, 10, 10, 10, 30]

        row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = col_widths[col_idx - 1]

        row = 5
        for cat in bank_result.get("categories", []):
            for rate in cat.get("rates", []):
                for col_idx, value in enumerate(
                    [
                        cat.get("category_name", ""),
                        cat.get("amount_slab", ""),
                        cat.get("scheme_name", ""),
                        rate.get("tenor_description", ""),
                        rate.get("min_days"),
                        rate.get("max_days"),
                        rate.get("rate_percent"),
                        rate.get("additional_info", ""),
                    ],
                    1,
                ):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if (row - 5) % 2 == 1:
                        cell.fill = alt_fill
                row += 1

        if row > 5:
            ws.auto_filter.ref = f"A4:H{row - 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()

    timestamp = datetime.now(timezone.utc)
    ts_str = timestamp.strftime("%Y%m%d_%H%M%S")

    _save_local_result(f"fd_rates_{ts_str}.xlsx", excel_bytes)
    _save_local_result("latest.xlsx", excel_bytes)

    # Upload Excel to Azure Blob Storage
    _upload_to_blob(
        f"fd_rates_{ts_str}.xlsx",
        excel_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    _upload_to_blob(
        "latest.xlsx",
        excel_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return jsonify(
        {
            "message": "Excel exported successfully",
            "blob_name": f"fd_rates_{ts_str}.xlsx",
            "latest_blob": "latest.xlsx",
            "bank_count": len(results),
            "exported_at": timestamp.isoformat(),
        }
    )


if __name__ == "__main__":
    print("\n  FD Rate Aggregator — Dev Server")
    print("  http://localhost:7071\n")
    app.run(host="0.0.0.0", port=7071, debug=True)
