"""
FD Rate Scraper — Azure Functions (Python v2 programming model)
HTTP-triggered functions for URL management, scraping, and export.
"""

import json
import logging
import os
import uuid
from datetime import datetime, timezone

import azure.functions as func
from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient

app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
URLS_FILE = os.path.join(os.path.dirname(__file__), "urls.json")
BLOB_CONTAINER = os.environ.get("BLOB_CONTAINER_NAME", "fd-rates")


def _get_blob_service_client() -> BlobServiceClient:
    account_name = os.environ["STORAGE_ACCOUNT_NAME"]
    account_url = f"https://{account_name}.blob.core.windows.net"
    return BlobServiceClient(account_url, credential=DefaultAzureCredential())


def _load_urls() -> list[dict]:
    if not os.path.exists(URLS_FILE):
        return []
    with open(URLS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_urls(urls: list[dict]) -> None:
    with open(URLS_FILE, "w", encoding="utf-8") as f:
        json.dump(urls, f, indent=2, ensure_ascii=False)


def _json_response(body, status_code=200):
    return func.HttpResponse(
        json.dumps(body, ensure_ascii=False, default=str),
        status_code=status_code,
        mimetype="application/json",
    )


# ---------------------------------------------------------------------------
# GET /api/urls — List all stored bank URLs
# ---------------------------------------------------------------------------
@app.route(route="urls", methods=["GET"])
def list_urls(req: func.HttpRequest) -> func.HttpResponse:
    return _json_response(_load_urls())


# ---------------------------------------------------------------------------
# POST /api/urls — Add a new bank URL
# ---------------------------------------------------------------------------
@app.route(route="urls", methods=["POST"])
def add_url(req: func.HttpRequest) -> func.HttpResponse:
    try:
        body = req.get_json()
    except ValueError:
        return _json_response({"error": "Invalid JSON body"}, 400)

    url = body.get("url", "").strip()
    bank_name = body.get("bank_name", "").strip()

    if not url or not bank_name:
        return _json_response({"error": "Both 'url' and 'bank_name' are required"}, 400)

    urls = _load_urls()
    new_entry = {
        "id": str(uuid.uuid4()),
        "url": url,
        "bank_name": bank_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    urls.append(new_entry)
    _save_urls(urls)

    return _json_response(new_entry, 201)


# ---------------------------------------------------------------------------
# DELETE /api/urls/{id} — Remove a URL by ID
# ---------------------------------------------------------------------------
@app.route(route="urls/{id}", methods=["DELETE"])
def delete_url(req: func.HttpRequest) -> func.HttpResponse:
    url_id = req.route_params.get("id")
    urls = _load_urls()
    original_len = len(urls)
    urls = [u for u in urls if u["id"] != url_id]

    if len(urls) == original_len:
        return _json_response({"error": "URL not found"}, 404)

    _save_urls(urls)
    return _json_response({"message": "Deleted", "id": url_id})


# ---------------------------------------------------------------------------
# POST /api/scrape — Trigger the Foundry Agent to scrape all stored URLs
# ---------------------------------------------------------------------------
@app.route(route="scrape", methods=["POST"])
def scrape_all(req: func.HttpRequest) -> func.HttpResponse:
    from agent.fd_rate_agent import scrape_all_urls

    urls = _load_urls()
    if not urls:
        return _json_response({"error": "No URLs configured. Add URLs first."}, 400)

    # Optional: filter by selected URL ids in request body. Default = scrape all.
    try:
        body = req.get_json()
    except ValueError:
        body = None
    selected_ids = (body or {}).get("ids") if isinstance(body, dict) else None
    if isinstance(selected_ids, list) and len(selected_ids) > 0:
        wanted = {str(i) for i in selected_ids}
        urls = [u for u in urls if str(u.get("id")) in wanted]
        if not urls:
            return _json_response(
                {"error": "None of the selected URL ids match configured URLs."}, 400
            )

    logger.info("Starting fetch for %d URLs", len(urls))
    import time as _time

    _t0 = _time.monotonic()
    scrape_output = scrape_all_urls(urls)
    elapsed_seconds = round(_time.monotonic() - _t0, 1)
    results = (
        scrape_output.get("results", scrape_output)
        if isinstance(scrape_output, dict)
        else scrape_output
    )
    token_usage = (
        scrape_output.get("token_usage", {}) if isinstance(scrape_output, dict) else {}
    )
    di_pages = (
        scrape_output.get("di_pages", 0) if isinstance(scrape_output, dict) else 0
    )

    # Build final payload
    timestamp = datetime.now(timezone.utc)
    payload = {
        "scraped_at": timestamp.isoformat(),
        "bank_count": len(results),
        "token_usage": token_usage,
        "di_pages": di_pages,
        "elapsed_seconds": elapsed_seconds,
        "results": results,
    }
    payload_json = json.dumps(payload, ensure_ascii=False, default=str)

    # Upload to Blob Storage
    blob_service = _get_blob_service_client()
    container = blob_service.get_container_client(BLOB_CONTAINER)

    # Ensure container exists
    try:
        container.create_container()
    except Exception:
        pass  # already exists

    ts_str = timestamp.strftime("%Y%m%d_%H%M%S")
    # Upload timestamped version
    container.upload_blob(
        name=f"fd_rates_{ts_str}.json",
        data=payload_json,
        overwrite=True,
    )
    # Upload latest.json
    container.upload_blob(
        name="latest.json",
        data=payload_json,
        overwrite=True,
    )

    logger.info("Fetch complete. Results uploaded to Blob Storage.")
    return _json_response(payload)


# ---------------------------------------------------------------------------
# GET /api/results/latest — Return the latest scraped results
# ---------------------------------------------------------------------------
@app.route(route="results/latest", methods=["GET"])
def get_latest_results(req: func.HttpRequest) -> func.HttpResponse:
    try:
        blob_service = _get_blob_service_client()
        container = blob_service.get_container_client(BLOB_CONTAINER)
        blob = container.download_blob("latest.json")
        data = json.loads(blob.readall())
        return _json_response(data)
    except Exception as e:
        logger.error("Failed to fetch latest results: %s", e)
        return _json_response({"error": "No results found. Run a scrape first."}, 404)


# ---------------------------------------------------------------------------
# POST /api/export-excel — Generate styled Excel and upload to Blob
# ---------------------------------------------------------------------------
@app.route(route="export-excel", methods=["POST"])
def export_excel(req: func.HttpRequest) -> func.HttpResponse:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Fetch latest results
    try:
        blob_service = _get_blob_service_client()
        container = blob_service.get_container_client(BLOB_CONTAINER)
        blob = container.download_blob("latest.json")
        data = json.loads(blob.readall())
    except Exception as e:
        return _json_response({"error": f"No results found to export: {e}"}, 404)

    results = data.get("results", [])
    if not results:
        return _json_response({"error": "No bank results to export"}, 400)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for bank_result in results:
        bank_name = bank_result.get("bank_name", "Unknown")
        # Excel sheet titles cannot contain: \ / ? * [ ] : and must be <= 31 chars
        safe_title = (
            "".join(c if c not in "\\/?*[]:" else "-" for c in bank_name)[:31].strip()
            or "Unknown"
        )
        base_title = safe_title
        suffix = 2
        while safe_title in wb.sheetnames:
            safe_title = f"{base_title[:28]}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(title=safe_title)

        # Title row
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"{bank_name} — Fixed Deposit Rates"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")

        # Effective date
        eff_date = bank_result.get("effective_date", "N/A")
        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Effective Date: {eff_date}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        if "error" in bank_result:
            ws["A4"].value = f"Error: {bank_result.get('error')}"
            ws["A5"].value = f"Reason: {bank_result.get('reason', 'N/A')}"
            continue

        # Headers
        headers = [
            "Category",
            "Amount Slab",
            "Scheme",
            "Tenor",
            "Min Days",
            "Max Days",
            "Rate (%)",
            "Additional Info",
        ]
        col_widths = [20, 20, 18, 25, 10, 10, 10, 30]

        row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = col_widths[col_idx - 1]

        row = 5
        categories = bank_result.get("categories", [])
        for cat in categories:
            cat_name = cat.get("category_name", "")
            amount_slab = cat.get("amount_slab", "")
            scheme = cat.get("scheme_name", "")
            rates = cat.get("rates", [])

            for rate_idx, rate in enumerate(rates):
                for col_idx, value in enumerate(
                    [
                        cat_name,
                        amount_slab,
                        scheme,
                        rate.get("tenor_description", ""),
                        rate.get("min_days"),
                        rate.get("max_days"),
                        rate.get("rate_percent"),
                        rate.get("additional_info", ""),
                    ],
                    1,
                ):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if (row - 5) % 2 == 1:
                        cell.fill = alt_fill
                row += 1

        # Auto-filter
        if row > 5:
            ws.auto_filter.ref = f"A4:H{row - 1}"

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()

    # Upload to Blob
    timestamp = datetime.now(timezone.utc)
    ts_str = timestamp.strftime("%Y%m%d_%H%M%S")

    container.upload_blob(
        name=f"fd_rates_{ts_str}.xlsx",
        data=excel_bytes,
        overwrite=True,
    )
    container.upload_blob(
        name="latest.xlsx",
        data=excel_bytes,
        overwrite=True,
    )

    return _json_response(
        {
            "message": "Excel exported successfully",
            "blob_name": f"fd_rates_{ts_str}.xlsx",
            "latest_blob": "latest.xlsx",
            "bank_count": len(results),
            "exported_at": timestamp.isoformat(),
        }
    )
